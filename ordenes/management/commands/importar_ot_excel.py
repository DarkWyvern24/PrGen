from pathlib import Path
from datetime import datetime, date

from django.core.management.base import BaseCommand
from ordenes.models import OrdenTrabajo
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa órdenes de trabajo desde bd.xlsx ubicado en esta misma carpeta"

    def handle(self, *args, **options):
        archivo_excel = Path(__file__).resolve().parent / "bd.xlsx"

        if not archivo_excel.exists():
            self.stdout.write(self.style.ERROR(
                f"No se encontró el archivo: {archivo_excel}"
            ))
            return

        try:
            wb = load_workbook(archivo_excel, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(
                f"Error al abrir el archivo Excel: {e}"
            ))
            return

        encabezados = self.obtener_encabezados(ws)
        if not encabezados:
            self.stdout.write(self.style.ERROR(
                "No se encontraron encabezados válidos en el Excel."
            ))
            return

        creadas = 0
        actualizadas = 0
        omitidas = 0

        fila_inicio_datos = encabezados["fila"] + 1

        for fila in range(fila_inicio_datos, ws.max_row + 1):
            numero = self.valor_celda(ws, fila, encabezados["columnas"].get("Numero"))

            if numero in (None, "", "None"):
                continue

            try:
                numero = int(numero)
            except Exception:
                self.stdout.write(self.style.WARNING(
                    f"Fila {fila}: Numero inválido '{numero}', se omite."
                ))
                omitidas += 1
                continue

            fecha = self.parsear_fecha(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Fecha"))
            )

            numero_cotizacion = self.parsear_entero(
                self.valor_celda(ws, fila, encabezados["columnas"].get("N° Cotizacion"))
            )

            cliente = self.limpiar_texto(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Cliente"))
            )
            referencia = self.limpiar_texto(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Referencia"))
            )
            atencion = self.limpiar_texto(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Atencion"))
            )
            responsable = self.limpiar_texto(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Responsable"))
            )
            estado = self.limpiar_texto(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Estado"))
            )
            nivel_urgencia = self.parsear_entero(
                self.valor_celda(ws, fila, encabezados["columnas"].get("Nivel de Urgencia"))
            )

            fecha_entrega = None
            if "Fecha Entrega" in encabezados["columnas"]:
                fecha_entrega = self.parsear_fecha(
                    self.valor_celda(ws, fila, encabezados["columnas"].get("Fecha Entrega"))
                )

            datos = {
                "fecha": fecha,
                "numero_cotizacion": numero_cotizacion,
                "cliente": cliente,
                "referencia": referencia,
                "atencion": atencion,
                "responsable": responsable,
                "estado": estado,
                "nivel_urgencia": nivel_urgencia,
            }

            if hasattr(OrdenTrabajo, "fecha_entrega"):
                datos["fecha_entrega"] = fecha_entrega

            try:
                obj, creado = OrdenTrabajo.objects.update_or_create(
                    numero=numero,
                    defaults=datos
                )

                if creado:
                    creadas += 1
                else:
                    actualizadas += 1

            except Exception as e:
                omitidas += 1
                self.stdout.write(self.style.WARNING(
                    f"Fila {fila}: error al guardar OT {numero}: {e}"
                ))

        self.stdout.write(self.style.SUCCESS(
            f"Importación finalizada. Creadas: {creadas}, actualizadas: {actualizadas}, omitidas: {omitidas}"
        ))

    def obtener_encabezados(self, ws):
        encabezados_esperados = {
            "Fecha",
            "Numero",
            "N° Cotizacion",
            "Cliente",
            "Referencia",
            "Atencion",
            "Responsable",
            "Estado",
            "Nivel de Urgencia",
        }

        aliases = {
            "Nº Cotizacion": "N° Cotizacion",
            "No Cotizacion": "N° Cotizacion",
            "Nro Cotizacion": "N° Cotizacion",
            "Nivel de urgencia": "Nivel de Urgencia",
            "Fecha Entrega": "Fecha Entrega",
        }

        for fila in range(1, ws.max_row + 1):
            columnas = {}
            encontrados = set()

            for col in range(1, ws.max_column + 1):
                valor = ws.cell(row=fila, column=col).value
                if valor is None:
                    continue

                texto = str(valor).strip()
                texto = aliases.get(texto, texto)

                if texto in encabezados_esperados or texto == "Fecha Entrega":
                    columnas[texto] = col
                    encontrados.add(texto)

            claves_minimas = {"Fecha", "Numero", "Cliente", "Estado"}
            if claves_minimas.issubset(encontrados):
                return {
                    "fila": fila,
                    "columnas": columnas,
                }

        return None

    def valor_celda(self, ws, fila, col):
        if not col:
            return None
        return ws.cell(row=fila, column=col).value

    def limpiar_texto(self, valor):
        if valor is None:
            return ""
        return str(valor).strip()

    def parsear_entero(self, valor):
        if valor in (None, ""):
            return None
        try:
            return int(valor)
        except Exception:
            try:
                return int(float(valor))
            except Exception:
                return None

    def parsear_fecha(self, valor):
        if valor in (None, ""):
            return None

        if isinstance(valor, datetime):
            return valor.date()

        if isinstance(valor, date):
            return valor

        if isinstance(valor, str):
            valor = valor.strip()

            formatos = [
                "%d-%m-%Y",
                "%d/%m/%Y",
                "%Y-%m-%d",
                "%d-%m-%y",
                "%d/%m/%y",
            ]

            for formato in formatos:
                try:
                    return datetime.strptime(valor, formato).date()
                except ValueError:
                    continue

        return None