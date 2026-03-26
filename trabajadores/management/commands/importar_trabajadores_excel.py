from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from trabajadores.models import Trabajador


class Command(BaseCommand):
    help = "Sincroniza trabajadores desde nomina.xlsx"

    def handle(self, *args, **options):
        archivo_excel = Path(__file__).resolve().parent / "nomina.xlsx"

        if not archivo_excel.exists():
            self.stdout.write(self.style.ERROR(
                f"No se encontró el archivo: {archivo_excel}"
            ))
            return

        try:
            wb = load_workbook(archivo_excel, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(
                f"Error al abrir el Excel: {e}"
            ))
            return

        creados = 0
        actualizados = 0
        omitidos = 0

        # CLAVE: guardar todos los RUT del Excel
        ruts_excel = set()

        for fila in range(5, ws.max_row + 1):
            nombre_raw = ws.cell(row=fila, column=2).value
            rut_raw = ws.cell(row=fila, column=3).value
            cargo_raw = ws.cell(row=fila, column=4).value

            if not nombre_raw and not rut_raw:
                continue

            nombre = self.normalizar_nombre(nombre_raw)
            rut = self.limpiar_texto(rut_raw)
            cargo = self.limpiar_texto(cargo_raw)

            if not nombre:
                omitidos += 1
                continue

            if rut:
                ruts_excel.add(rut)

                trabajador, creado = Trabajador.objects.update_or_create(
                    rut=rut,
                    defaults={
                        "nombre": nombre,
                        "cargo": cargo,
                        "activo": True,
                    }
                )
            else:
                trabajador, creado = Trabajador.objects.update_or_create(
                    nombre=nombre,
                    defaults={
                        "cargo": cargo,
                        "activo": True,
                    }
                )

            if creado:
                creados += 1
            else:
                actualizados += 1

        # INACTIVAR LOS QUE NO ESTÁN EN EXCEL
        desactivados = 0

        trabajadores_db = Trabajador.objects.all()

        for trabajador in trabajadores_db:
            if trabajador.rut and trabajador.rut not in ruts_excel:
                if trabajador.activo:
                    trabajador.activo = False
                    trabajador.save()
                    desactivados += 1

        self.stdout.write(self.style.SUCCESS(
            f"""
Sincronización completada:
- Creados: {creados}
- Actualizados: {actualizados}
- Inactivados: {desactivados}
- Omitidos: {omitidos}
"""
        ))

    def limpiar_texto(self, valor):
        if valor is None:
            return ""
        return str(valor).strip()

    def normalizar_nombre(self, valor):
        texto = self.limpiar_texto(valor)
        if not texto:
            return ""

        partes = [p for p in texto.split() if p.strip()]
        if len(partes) < 3:
            return texto.title()

        apellidos = partes[:2]
        nombres = partes[2:]

        nombre_final = " ".join(nombres + apellidos)
        return nombre_final.title()